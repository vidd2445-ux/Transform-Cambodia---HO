from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import json, os, csv, io, random, string, base64, re, hashlib, urllib.request, math, math
import qrcode
import smtplib
import threading, time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from PIL import Image as PILImage
from datetime import datetime, date, timedelta
from functools import wraps
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    raise RuntimeError('SECRET_KEY must be configured in the environment.')

DATA_DIR = os.environ.get('DATA_DIR', os.path.join(os.path.dirname(__file__), 'data'))
STAFF_FILE = os.path.join(DATA_DIR, 'staff.json')
CLASSES_FILE = os.path.join(DATA_DIR, 'classes.json')
ATTENDANCE_FILE = os.path.join(DATA_DIR, 'attendance.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
OTP_FILE = os.path.join(DATA_DIR, 'otp_codes.json')

# ─── Telegram Bot Configuration ───
TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '')
TELEGRAM_BOT_USERNAME = os.environ.get('TELEGRAM_BOT_USERNAME', 'TransformCambodiastaff_bot')

# ─── SMTP Configuration (legacy fallback) ───
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER)

# ─── Helper Functions ───
def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            return json.load(f)
    return []

def save_json(filepath, data):
    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)

def init_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(STAFF_FILE):
        save_json(STAFF_FILE, [])
    if not os.path.exists(CLASSES_FILE):
        save_json(CLASSES_FILE, [
            {"id": "C001", "name": "Transform Cambodia - HO", "section": "Stueng Mean Chey"}
        ])
    if not os.path.exists(ATTENDANCE_FILE):
        save_json(ATTENDANCE_FILE, [])
    if not os.path.exists(USERS_FILE):
        save_json(USERS_FILE, [
            {"id": "U001", "username": "admin", "password": "admin123", "role": "admin", "name": "Administrator", "phone": "+855 96 123 4567", "email": "admin@transformcambodia.org", "telegram_chat_id": ""}
        ])
    if not os.path.exists(OTP_FILE):
        save_json(OTP_FILE, [])

def generate_id(prefix, data):
    max_num = 0
    for item in data:
        if item.get('id', '').startswith(prefix):
            try:
                num = int(item['id'].replace(prefix, ''))
                max_num = max(max_num, num)
            except:
                pass
    return f"{prefix}{max_num + 1:03d}"

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

def send_telegram_otp(chat_id, otp_code):
    """Send a 6-digit OTP code to the user via Telegram Bot."""
    if not TELEGRAM_BOT_TOKEN or not chat_id:
        return False, 'Telegram not configured or no chat ID. OTP shown above for development.'
    try:
        message = (
            "🔐 *Transform Cambodia - Password Reset*\n\n"
            f"Your verification code is: `{otp_code}`\n\n"
            "⏱ This code expires in 10 minutes.\n\n"
            "If you did not request this, please ignore this message."
        )
        url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage'
        payload = json.dumps({
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'Markdown'
        }).encode('utf-8')
        req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode())
        if result.get('ok'):
            return True, 'Verification code sent via Telegram. Check your messages.'
        else:
            desc = result.get('description', 'Unknown error')
            return False, f'Telegram error: {desc}'
    except Exception as e:
        return False, f'Failed to send Telegram message: {str(e)}'

def send_email_otp(to_email, otp_code):
    """Legacy fallback: Send OTP via email (SMTP)."""
    if not SMTP_USER or not SMTP_PASSWORD:
        return False, 'SMTP not configured. OTP shown above for development.'
    try:
        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        msg['Subject'] = 'Transform Cambodia - Password Reset Verification Code'
        body = (
            f"Hello,\n\n"
            f"Your verification code for password reset is: {otp_code}\n\n"
            f"This code expires in 10 minutes.\n\n"
            f"If you did not request this, please ignore this email.\n\n"
            f"- Transform Cambodia Attendance System"
        )
        msg.attach(MIMEText(body, 'plain'))
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(SMTP_FROM, to_email, msg.as_string())
        server.quit()
        return True, f'Verification code sent to {to_email}. Check your inbox.'
    except Exception as e:
        return False, f'Failed to send email: {str(e)}'

def normalize_cambodia_phone(phone):
    """Normalize Cambodia phone number for comparison."""
    if not phone:
        return ''
    phone = re.sub(r'[\s\-\.\(\)]+', '', phone.strip())
    if phone.startswith('+855'):
        phone = '0' + phone[4:]
    elif phone.startswith('855'):
        phone = '0' + phone[3:]
    if not phone.startswith('0'):
        phone = '0' + phone
    return phone

# ─── IP / Location Helper ───
def get_client_ip():
    forwarded = request.headers.get('X-Forwarded-For', '')
    if forwarded:
        return forwarded.split(',')[0].strip()
    real_ip = request.headers.get('X-Real-IP', '')
    if real_ip:
        return real_ip.strip()
    return request.remote_addr or '127.0.0.1'

def get_location_from_ip(ip):
    if ip in ('127.0.0.1', 'localhost', '::1') or ip.startswith('192.168.') or ip.startswith('10.') or ip.startswith('172.'):
        return {'city': 'Local', 'region': 'Local', 'country': 'Local', 'lat': '', 'lon': ''}
    try:
        url = f'http://ip-api.com/json/{ip}?fields=status,country,regionName,city,lat,lon'
        req = urllib.request.Request(url, headers={'User-Agent': 'TransformCambodiaAttendance/1.0'})
        with urllib.request.urlopen(req, timeout=3) as resp:
            data = json.loads(resp.read().decode())
        if data.get('status') == 'success':
            return {
                'city': data.get('city', ''),
                'region': data.get('regionName', ''),
                'country': data.get('country', ''),
                'lat': data.get('lat', ''),
                'lon': data.get('lon', '')
            }
    except Exception:
        pass
    return {'city': '', 'region': '', 'country': '', 'lat': '', 'lon': ''}

def get_client_location_info():
    ip = get_client_ip()
    loc = get_location_from_ip(ip)
    return {
        'ip_address': ip,
        'ip_city': loc.get('city', ''),
        'ip_region': loc.get('region', ''),
        'ip_country': loc.get('country', '')
    }

# ─── GPS Geofence Helpers ───
def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate distance in meters between two GPS points using the Haversine formula."""
    R = 6371000  # Earth radius in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

def check_geofence(gps_lat, gps_lon, class_id):
    """Check if GPS coords are within the geofence radius of the given Location.
    Returns (is_within, distance_m, centre_name, centre_lat, centre_lon, radius_m)."""
    if not gps_lat or not gps_lon:
        return None, None, None, None, None, None  # GPS not provided
    try:
        gps_lat = float(gps_lat)
        gps_lon = float(gps_lon)
    except (ValueError, TypeError):
        return None, None, None, None, None, None
    classes = load_json(CLASSES_FILE)
    centre = next((c for c in classes if c['id'] == class_id), None)
    if not centre or 'lat' not in centre or 'lon' not in centre:
        return None, None, None, None, None, None  # Centre has no GPS coords
    centre_lat = float(centre['lat'])
    centre_lon = float(centre['lon'])
    radius_m = float(centre.get('radius_meters', 200))
    dist = haversine_distance(gps_lat, gps_lon, centre_lat, centre_lon)
    is_within = dist <= radius_m
    return is_within, round(dist, 1), centre['name'], centre_lat, centre_lon, radius_m

# ─── Google Sign-In Configuration ───
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
GOOGLE_DISCOVERY_URL = 'https://accounts.google.com/.well-known/openid-configuration'

def make_excel(headers, rows, title="Export"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    gold_fill = PatternFill(start_color="D4A843", end_color="D4A843", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), max((len(str(r[col_idx-1])) for r in rows), default=0))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)
    return wb

init_data()

# ─── Telegram Bot Polling (works on localhost without public URL) ───
telegram_polling_offset = 0
telegram_polling_active = True

def telegram_poll_loop():
    """Background thread that polls Telegram getUpdates to process /start commands.
    This makes the bot work on localhost or any server without a public webhook URL."""
    global telegram_polling_offset, telegram_polling_active
    if not TELEGRAM_BOT_TOKEN:
        return
    # Delete any existing webhook first so polling works
    try:
        url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/deleteWebhook'
        payload = json.dumps({'drop_pending_updates': False}).encode('utf-8')
        req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            app.logger.info('Telegram: deleted existing webhook for polling mode')
    except Exception:
        pass
    time.sleep(2)
    while telegram_polling_active:
        try:
            url = (f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates'
                   f'?offset={telegram_polling_offset}&timeout=30&allowed_updates=["message"]')
            req = urllib.request.Request(url)
            with urllib.request.urlopen(req, timeout=35) as resp:
                result = json.loads(resp.read().decode())
            if result.get('ok') and result.get('result'):
                for update in result['result']:
                    telegram_polling_offset = update['update_id'] + 1
                    message = update.get('message', {})
                    if message:
                        _process_telegram_update(message)
        except Exception as e:
            time.sleep(3)
        time.sleep(0.5)

def _process_telegram_update(message):
    """Process an incoming Telegram message (shared logic for both webhook and polling)."""
    chat_id = str(message.get('chat', {}).get('id', ''))
    text = (message.get('text') or '').strip()
    from_user = message.get('from', {})
    telegram_username = (from_user.get('username') or '').strip()
    telegram_first = (from_user.get('first_name') or '').strip()

    if not (text.startswith('/start') and chat_id):
        return

    parts = text.split()
    lookup_key = parts[1].strip().lstrip('@') if len(parts) > 1 else ''

    users = load_json(USERS_FILE)
    matched_user = None

    # Check if lookup_key is a link token
    link_tokens_file = os.path.join(DATA_DIR, 'telegram_link_tokens.json')
    link_tokens = load_json(link_tokens_file) if os.path.exists(link_tokens_file) else []
    if lookup_key and link_tokens:
        token_entry = next((t for t in link_tokens if t['token'] == lookup_key), None)
        if token_entry:
            expires = datetime.fromisoformat(token_entry['expires_at'])
            if datetime.now() > expires:
                reply = (
                    "❌ *Link Token Expired*\n\n"
                    "The linking token has expired (valid for 30 minutes).\n"
                    "Please go back to the Forgot Password page and generate a new one."
                )
            else:
                matched_user = next((u for u in users if u['id'] == token_entry['user_id']), None)
                if matched_user:
                    matched_user['telegram_chat_id'] = chat_id
                    save_json(USERS_FILE, users)
                    link_tokens = [t for t in link_tokens if t['token'] != lookup_key]
                    save_json(link_tokens_file, link_tokens)
                    reply = (
                        "✅ *Account Linked Successfully!*\n\n"
                        f"Your Telegram is now connected to the Transform Cambodia account: *{matched_user['username']}*\n\n"
                        "You will receive password reset OTPs here."
                    )
                else:
                    reply = "❌ *User Not Found*\n\nThe account associated with this token could not be found."
            _send_telegram_reply(chat_id, reply)
            return

    # Fallback: match by username or email
    if lookup_key and not matched_user:
        matched_user = next((u for u in users if u.get('username', '').lower() == lookup_key.lower()), None)
        if not matched_user:
            matched_user = next((u for u in users if u.get('email', '').lower() == lookup_key.lower()), None)

    # Try matching by Telegram username
    if not matched_user and telegram_username:
        matched_user = next((u for u in users if u.get('username', '').lower() == telegram_username.lower()), None)

    if matched_user:
        matched_user['telegram_chat_id'] = chat_id
        save_json(USERS_FILE, users)
        reply = (
            "✅ *Account Linked Successfully!*\n\n"
            f"Your Telegram is now connected to the Transform Cambodia account: *{matched_user['username']}*\n\n"
            "You will receive password reset OTPs here."
        )
    else:
        reply = (
            f"👋 Hello {telegram_first or 'there'}!\n\n"
            "To link your Telegram to your Transform Cambodia account, please send:\n"
            "`/start YOUR_USERNAME`\n\n"
            "Replace *YOUR_USERNAME* with the username you use to sign in.\n"
            "Example: `/start sreypi`"
        )
    _send_telegram_reply(chat_id, reply)

def _send_telegram_reply(chat_id, text):
    """Send a reply message via Telegram Bot API."""
    if not TELEGRAM_BOT_TOKEN:
        return
    try:
        url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage'
        payload = json.dumps({
            'chat_id': chat_id,
            'text': text,
            'parse_mode': 'Markdown'
        }).encode('utf-8')
        req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            pass
    except Exception:
        pass

# Start Telegram polling thread (daemon so it stops when main process exits)
if TELEGRAM_BOT_TOKEN:
    _poll_thread = threading.Thread(target=telegram_poll_loop, daemon=True)
    _poll_thread.start()

# ─── Auth Decorators ───
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = request.cookies.get('user_id')
        if not user:
            flash('Please login first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = request.cookies.get('user_id')
        users = load_json(USERS_FILE)
        user = next((u for u in users if u['id'] == user_id), None)
        if not user or user['role'] != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# ─── Auth Routes ───
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        users = load_json(USERS_FILE)
        user = next((u for u in users if u['username'] == username and u['password'] == password), None)
        if user:
            resp = redirect(url_for('dashboard'))
            resp.set_cookie('user_id', user['id'])
            resp.set_cookie('user_role', user['role'])
            resp.set_cookie('user_name', user['name'])
            flash(f'Welcome, {user["name"]}!', 'success')
            return resp
        flash('Invalid username or password.', 'danger')
    return render_template('login.html', google_enabled=bool(GOOGLE_CLIENT_ID))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        name = username  # username serves as full name
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        role = 'staff'  # staff only — no admin option
        if not all([username, password, email]):
            flash('Username, password, and email are required.', 'danger')
            return render_template('signup.html', google_enabled=bool(GOOGLE_CLIENT_ID))
        if password != confirm:
            flash('Passwords do not match.', 'danger')
            return render_template('signup.html', google_enabled=bool(GOOGLE_CLIENT_ID))
        if len(password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('signup.html', google_enabled=bool(GOOGLE_CLIENT_ID))
        users = load_json(USERS_FILE)
        if any(u['username'] == username for u in users):
            flash('Username already taken.', 'danger')
            return render_template('signup.html', google_enabled=bool(GOOGLE_CLIENT_ID))
        if any(u.get('email', '').lower() == email.lower() for u in users):
            flash('Email already registered.', 'danger')
            return render_template('signup.html', google_enabled=bool(GOOGLE_CLIENT_ID))
        # Auto-assign to Transform Cambodia - HO
        location = 'Transform Cambodia - HO'
        classes_list = load_json(CLASSES_FILE)
        assigned_class = next((c for c in classes_list if c['name'] == location), None)
        class_id = assigned_class['id'] if assigned_class else 'C001'

        new_user = {
            'id': generate_id('U', users),
            'username': username,
            'password': password,
            'role': role,
            'name': name,
            'email': email,
            'phone': phone,
            'telegram_chat_id': ''
        }
        users.append(new_user)
        save_json(USERS_FILE, users)

        # Auto-add to staff.json with Transform Cambodia - HO
        staff_list = load_json(STAFF_FILE)
        new_staff = {
            'id': generate_id('S', staff_list),
            'name': name,
            'class_id': class_id,
            'class_name': location,
            'phone': phone,
            'status': 'active',
            'roll_no': ''
        }
        staff_list.append(new_staff)
        save_json(STAFF_FILE, staff_list)

        flash('Account created successfully! You have been assigned to Transform Cambodia - HO. Please sign in.', 'success')
        return redirect(url_for('login'))
    return render_template('signup.html', google_enabled=bool(GOOGLE_CLIENT_ID))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST' and request.form.get('step') == 'send_code':
        email = request.form.get('email', '').strip()
        users = load_json(USERS_FILE)
        user = next((u for u in users if u.get('email', '').lower() == email.lower()), None)
        if not user:
            flash('Email address not found. Please check and try again.', 'danger')
            return render_template('forgot_password.html', step='send')
        # Check Telegram linked BEFORE generating OTP — no website/email fallback
        chat_id = user.get('telegram_chat_id', '')
        if not chat_id:
            flash('You must link your Telegram account before you can reset your password. OTP codes are only delivered via Telegram for security.', 'warning')
            return redirect(url_for('link_telegram'))

        # Generate and store OTP only after confirming Telegram is linked
        otp = generate_otp()
        otp_list = load_json(OTP_FILE)
        otp_list = [o for o in otp_list if o['user_id'] != user['id']]
        otp_list.append({
            'user_id': user['id'],
            'email': email,
            'code': otp,
            'created_at': datetime.now().isoformat(),
            'expires_at': (datetime.now() + timedelta(minutes=10)).isoformat()
        })
        save_json(OTP_FILE, otp_list)

        # Send via Telegram only — never display OTP on website
        sent, msg = send_telegram_otp(chat_id, otp)
        if sent:
            flash('Verification code sent to your Telegram. Check your messages.', 'info')
        else:
            flash('Failed to send verification code via Telegram. Please try again or contact an administrator.', 'danger')
        return render_template('forgot_password.html', step='verify', email=email, user_id=user['id'])
    if request.method == 'POST' and request.form.get('step') == 'verify_code':
        user_id = request.form.get('user_id')
        code = request.form.get('code', '').strip()
        otp_list = load_json(OTP_FILE)
        otp_entry = next((o for o in otp_list if o['user_id'] == user_id and o['code'] == code), None)
        if not otp_entry:
            flash('Invalid verification code.', 'danger')
            return render_template('forgot_password.html', step='verify', email=request.form.get('email',''), user_id=user_id)
        expires = datetime.fromisoformat(otp_entry['expires_at'])
        if datetime.now() > expires:
            flash('Verification code expired. Please request a new one.', 'danger')
            return render_template('forgot_password.html', step='send')
        return render_template('forgot_password.html', step='reset', user_id=user_id)
    if request.method == 'POST' and request.form.get('step') == 'reset_password':
        user_id = request.form.get('user_id')
        new_pass = request.form.get('new_password', '')
        confirm_pass = request.form.get('confirm_password', '')
        if new_pass != confirm_pass:
            flash('Passwords do not match.', 'danger')
            return render_template('forgot_password.html', step='reset', user_id=user_id)
        if len(new_pass) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('forgot_password.html', step='reset', user_id=user_id)
        users = load_json(USERS_FILE)
        for u in users:
            if u['id'] == user_id:
                u['password'] = new_pass
                break
        save_json(USERS_FILE, users)
        otp_list = load_json(OTP_FILE)
        otp_list = [o for o in otp_list if o['user_id'] != user_id]
        save_json(OTP_FILE, otp_list)
        flash('Password reset successfully! Please sign in.', 'success')
        return redirect(url_for('login'))
    return render_template('forgot_password.html', step='send')

# ─── Telegram Bot Webhook for Account Linking ───
@app.route('/telegram/webhook', methods=['GET', 'POST'])
def telegram_webhook():
    """Telegram bot webhook endpoint.
    GET: Verification ping (for setting up webhook).
    POST: Process incoming Telegram messages (e.g. /start command for account linking)."""
    if request.method == 'GET':
        return jsonify({'status': 'ok', 'message': 'Telegram webhook is active'}), 200

    data = request.get_json(silent=True) or {}
    message = data.get('message', {})
    if not message:
        return jsonify({'ok': True}), 200

    chat_id = str(message.get('chat', {}).get('id', ''))
    text = (message.get('text') or '').strip()
    from_user = message.get('from', {})
    telegram_username = (from_user.get('username') or '').strip()
    telegram_first = (from_user.get('first_name') or '').strip()

    if text.startswith('/start') and chat_id:
        # Extract optional parameter: /start username_or_link_token
        parts = text.split()
        lookup_key = parts[1].strip().lstrip('@') if len(parts) > 1 else ''

        users = load_json(USERS_FILE)
        matched_user = None

        # First: check if lookup_key is a link token
        link_tokens_file = os.path.join(DATA_DIR, 'telegram_link_tokens.json')
        link_tokens = load_json(link_tokens_file) if os.path.exists(link_tokens_file) else []
        if lookup_key and link_tokens:
            token_entry = next((t for t in link_tokens if t['token'] == lookup_key), None)
            if token_entry:
                # Verify token hasn't expired
                from datetime import timezone
                expires = datetime.fromisoformat(token_entry['expires_at'])
                if datetime.now() > expires:
                    reply = (
                        f"❌ *Link Token Expired*\n\n"
                        f"The linking token has expired (valid for 30 minutes).\n"
                        f"Please go back to the Forgot Password page and generate a new one."
                    )
                else:
                    matched_user = next((u for u in users if u['id'] == token_entry['user_id']), None)
                    if matched_user:
                        matched_user['telegram_chat_id'] = chat_id
                        save_json(USERS_FILE, users)
                        # Remove used token
                        link_tokens = [t for t in link_tokens if t['token'] != lookup_key]
                        save_json(link_tokens_file, link_tokens)
                        reply = (
                            f"✅ *Account Linked Successfully!*\n\n"
                            f"Your Telegram is now connected to the Transform Cambodia account: *{matched_user['username']}*\n\n"
                            f"You will receive password reset OTPs here."
                        )
                    else:
                        reply = (
                            f"❌ *User Not Found*\n\n"
                            f"The account associated with this token could not be found."
                        )
                # Send reply and return
                if TELEGRAM_BOT_TOKEN:
                    try:
                        url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage'
                        payload = json.dumps({
                            'chat_id': chat_id,
                            'text': reply,
                            'parse_mode': 'Markdown'
                        }).encode('utf-8')
                        req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
                        urllib.request.urlopen(req, timeout=10)
                    except Exception:
                        pass
                return jsonify({'ok': True}), 200

        # Fallback: try to match by Telegram username provided after /start
        if lookup_key and not matched_user:
            matched_user = next((u for u in users if u.get('username', '').lower() == lookup_key.lower()), None)
            if not matched_user:
                matched_user = next((u for u in users if u.get('email', '').lower() == lookup_key.lower()), None)

        # Try matching by Telegram username (if user set their app username same as Telegram)
        if not matched_user and telegram_username:
            matched_user = next((u for u in users if u.get('username', '').lower() == telegram_username.lower()), None)

        if matched_user:
            # Update the user's telegram_chat_id
            matched_user['telegram_chat_id'] = chat_id
            save_json(USERS_FILE, users)
            reply = (
                f"✅ *Account Linked Successfully!*\n\n"
                f"Your Telegram is now connected to the Transform Cambodia account: *{matched_user['username']}*\n\n"
                f"You will receive password reset OTPs here."
            )
        else:
            # No match found — ask user to provide their username
            reply = (
                f"👋 Hello {telegram_first or 'there'}!\n\n"
                f"To link your Telegram to your Transform Cambodia account, please send:\n"
                f"`/start YOUR_USERNAME`\n\n"
                f"Replace *YOUR_USERNAME* with the username you use to sign in.\n"
                f"Example: `/start john.doe`"
            )

        # Send reply via Telegram API
        if TELEGRAM_BOT_TOKEN:
            try:
                url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage'
                payload = json.dumps({
                    'chat_id': chat_id,
                    'text': reply,
                    'parse_mode': 'Markdown'
                }).encode('utf-8')
                req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
                urllib.request.urlopen(req, timeout=10)
            except Exception:
                pass

    return jsonify({'ok': True}), 200


@app.route('/setup-webhook', methods=['GET', 'POST'])
@admin_required
def setup_webhook():
    """Admin route to register the Telegram webhook for production HTTPS deployments.
    On localhost, polling is used automatically; in production with a public URL, use this
    to switch from polling to webhook mode."""
    global telegram_polling_active
    if request.method == 'POST':
        base_url = request.form.get('base_url', '').strip().rstrip('/')
        if not base_url:
            flash('Please provide the base URL of your public server.', 'danger')
            return render_template('setup_webhook.html')
        webhook_url = f'{base_url}/telegram/webhook'
        try:
            url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setWebhook'
            payload = json.dumps({'url': webhook_url, 'allowed_updates': ['message']}).encode('utf-8')
            req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                result = json.loads(resp.read().decode())
            if result.get('ok'):
                telegram_polling_active = False  # Stop polling since webhook is now active
                flash(f'Webhook set successfully! URL: {webhook_url}', 'success')
            else:
                flash(f'Failed to set webhook: {result.get("description", "Unknown error")}', 'danger')
        except Exception as e:
            flash(f'Error setting webhook: {str(e)}', 'danger')
    # Get current webhook info
    webhook_info = {}
    try:
        url = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getWebhookInfo'
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode())
        if result.get('ok'):
            webhook_info = result.get('result', {})
    except Exception:
        pass
    return render_template('setup_webhook.html', webhook_info=webhook_info)


@app.route('/link-telegram', methods=['GET', 'POST'])
def link_telegram():
    """Web page for linking a Telegram account to the user's account.
    This is accessed from the Forgot Password page when no telegram_chat_id is found."""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        users = load_json(USERS_FILE)
        user = next((u for u in users if u.get('username', '').lower() == username.lower()), None)
        if not user:
            # Also try email
            user = next((u for u in users if u.get('email', '').lower() == username.lower()), None)
        if not user:
            flash('Account not found. Please check your username or email.', 'danger')
            return render_template('link_telegram.html')
        # Generate a linking token
        link_token = hashlib.sha256(f"{user['id']}:{datetime.now().isoformat()}:{app.secret_key}".encode()).hexdigest()[:16]
        link_list = load_json(os.path.join(DATA_DIR, 'telegram_link_tokens.json')) if os.path.exists(os.path.join(DATA_DIR, 'telegram_link_tokens.json')) else []
        # Remove old tokens for this user
        link_list = [t for t in link_list if t['user_id'] != user['id']]
        link_list.append({
            'user_id': user['id'],
            'token': link_token,
            'username': user['username'],
            'created_at': datetime.now().isoformat(),
            'expires_at': (datetime.now() + timedelta(minutes=30)).isoformat()
        })
        save_json(os.path.join(DATA_DIR, 'telegram_link_tokens.json'), link_list)
        deep_link = f'https://t.me/{TELEGRAM_BOT_USERNAME}?start={link_token}'
        flash(f'Link request created! Click the button below to open Telegram and connect your account.', 'info')
        return render_template('link_telegram.html', deep_link=deep_link, username=user['username'], token_created=True)

    return render_template('link_telegram.html')

@app.route('/logout')
def logout():
    resp = redirect(url_for('login'))
    resp.set_cookie('user_id', '', expires=0)
    resp.set_cookie('user_role', '', expires=0)
    resp.set_cookie('user_name', '', expires=0)
    flash('Logged out successfully.', 'info')
    return resp

# ─── Dashboard ───
@app.route('/dashboard')
@login_required
def dashboard():
    staff = load_json(STAFF_FILE)
    classes = load_json(CLASSES_FILE)
    attendance = load_json(ATTENDANCE_FILE)
    today = date.today().isoformat()
    today_records = [r for r in attendance if r['date'] == today]
    total_staff = len(staff)
    total_centres = len(classes)
    present_today = sum(1 for r in today_records if r['status'] == 'present')
    absent_today = sum(1 for r in today_records if r['status'] == 'absent')
    late_today = sum(1 for r in today_records if r['status'] == 'late')
    # Time records today
    time_records = [r for r in today_records if r.get('time_in')]
    checked_in = sum(1 for r in time_records if r.get('time_in'))
    checked_out = sum(1 for r in time_records if r.get('time_out'))
    return render_template('dashboard.html',
        total_staff=total_staff, total_centres=total_centres,
        present_today=present_today, absent_today=absent_today, late_today=late_today,
        checked_in=checked_in, checked_out=checked_out,
        staff=staff, classes=classes, attendance=attendance, today=today)

# ─── Staff ───
@app.route('/staff')
@login_required
def staff():
    staff_list = load_json(STAFF_FILE)
    classes = load_json(CLASSES_FILE)
    for s in staff_list:
        cls = next((c for c in classes if c['id'] == s.get('class_id')), None)
        s['class_name'] = cls['name'] if cls else 'N/A'
    return render_template('staff.html', staff=staff_list, classes=classes)

@app.route('/staff/add', methods=['POST'])
@login_required
def add_staff():
    staff_list = load_json(STAFF_FILE)
    member = {
        'id': generate_id('S', staff_list),
        'name': request.form.get('name'),
        'class_id': request.form.get('class_id'),
        'phone': request.form.get('contact', ''),
        'admission_date': date.today().isoformat(),
        'status': 'active'
    }
    staff_list.append(member)
    save_json(STAFF_FILE, staff_list)
    flash('Staff enrolled successfully!', 'success')
    return redirect(url_for('staff'))

@app.route('/staff/edit', methods=['POST'])
@login_required
def edit_staff():
    sid = request.form.get('id')
    staff_list = load_json(STAFF_FILE)
    for s in staff_list:
        if s['id'] == sid:
            s['name'] = request.form.get('name', s['name'])
            s['class_id'] = request.form.get('class_id', s['class_id'])
            s['phone'] = request.form.get('contact', s.get('phone', ''))
            break
    save_json(STAFF_FILE, staff_list)
    flash('Staff record updated successfully!', 'success')
    return redirect(url_for('staff'))

@app.route('/staff/delete', methods=['POST'])
@login_required
@admin_required
def delete_staff():
    sid = request.form.get('id')
    staff_list = load_json(STAFF_FILE)
    staff_list = [s for s in staff_list if s['id'] != sid]
    save_json(STAFF_FILE, staff_list)
    flash('Staff removed successfully!', 'success')
    return redirect(url_for('staff'))

@app.route('/staff/export')
@login_required
def export_staff():
    fmt = request.args.get('format', 'csv')
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    headers = ['Location', 'Staff Name', 'Contact']
    rows = []
    for s in staff_list:
        cls = next((c for c in classes_list if c['id'] == s.get('class_id')), None)
        rows.append([
            cls['name'] if cls else 'N/A',
            s['name'],
            s.get('phone', '')
        ])
    if fmt == 'excel':
        wb = make_excel(headers, rows, 'Staff')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='staff_export.xlsx')
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv', as_attachment=True, download_name='staff_export.csv')

# ─── Locations ───
@app.route('/classes')
@login_required
def classes():
    classes_list = load_json(CLASSES_FILE)
    staff_list = load_json(STAFF_FILE)
    for c in classes_list:
        c['student_count'] = sum(1 for s in staff_list if s.get('class_id') == c['id'] and s.get('status') == 'active')
    return render_template('classes.html', classes=classes_list)

@app.route('/classes/add', methods=['POST'])
@login_required
@admin_required
def add_class():
    classes_list = load_json(CLASSES_FILE)
    cls = {
        'id': generate_id('C', classes_list),
        'name': request.form.get('name'),
        'section': request.form.get('section', 'Central')
    }
    classes_list.append(cls)
    save_json(CLASSES_FILE, classes_list)
    flash('Location added successfully!', 'success')
    return redirect(url_for('classes'))

@app.route('/classes/add-quick', methods=['POST'])
@login_required
@admin_required
def add_class_quick():
    classes_list = load_json(CLASSES_FILE)
    cls = {
        'id': generate_id('C', classes_list),
        'name': request.form.get('name'),
        'section': request.form.get('section', 'Central')
    }
    classes_list.append(cls)
    save_json(CLASSES_FILE, classes_list)
    return jsonify({'success': True, 'id': cls['id'], 'name': cls['name']})

@app.route('/classes/edit', methods=['POST'])
@login_required
@admin_required
def edit_class():
    cid = request.form.get('id')
    classes_list = load_json(CLASSES_FILE)
    for c in classes_list:
        if c['id'] == cid:
            c['name'] = request.form.get('name', c['name'])
            c['section'] = request.form.get('section', c.get('section', ''))
            break
    save_json(CLASSES_FILE, classes_list)
    flash('Location updated successfully!', 'success')
    return redirect(url_for('classes'))

@app.route('/classes/delete', methods=['POST'])
@login_required
@admin_required
def delete_class():
    cid = request.form.get('id')
    classes_list = load_json(CLASSES_FILE)
    classes_list = [c for c in classes_list if c['id'] != cid]
    staff_list = load_json(STAFF_FILE)
    staff_list = [s for s in staff_list if s.get('class_id') != cid]
    save_json(STAFF_FILE, staff_list)
    save_json(CLASSES_FILE, classes_list)
    flash('Location removed successfully!', 'success')
    return redirect(url_for('classes'))

# ─── Attendance ───
@app.route('/attendance/mark', methods=['GET', 'POST'])
@login_required
def mark_attendance():
    classes_list = load_json(CLASSES_FILE)
    staff_list = load_json(STAFF_FILE)
    today = date.today().isoformat()
    return render_template('mark_attendance.html',
        classes=classes_list, staff=staff_list, today=today)

@app.route('/attendance/save', methods=['POST'])
@login_required
def save_attendance():
    data = request.get_json()
    class_id = data.get('class_id')
    att_date = data.get('date')
    records = data.get('records', [])
    attendance_list = load_json(ATTENDANCE_FILE)
    # Remove existing records for this class+date
    attendance_list = [r for r in attendance_list
        if not (r['class_id'] == class_id and r['date'] == att_date)]
    loc_info = get_client_location_info()
    for rec in records:
        attendance_list.append({
            'id': generate_id('A', attendance_list),
            'student_id': rec['student_id'],
            'class_id': class_id,
            'date': att_date,
            'status': rec['status'],
            'time_in': rec.get('time_in', ''),
            'time_out': rec.get('time_out', ''),
            'marked_by': request.cookies.get('user_name', 'Unknown'),
            'marked_at': datetime.now().isoformat(),
            'ip_address': loc_info['ip_address'],
            'ip_city': loc_info['ip_city'],
            'ip_region': loc_info['ip_region'],
            'ip_country': loc_info['ip_country']
        })
    save_json(ATTENDANCE_FILE, attendance_list)
    return jsonify({'success': True})

# ─── QR Code Attendance ───
@app.route('/attendance/qr')
@login_required
def qr_attendance_page():
    classes_list = load_json(CLASSES_FILE)
    staff_list = load_json(STAFF_FILE)
    today = date.today().isoformat()
    attendance_list = load_json(ATTENDANCE_FILE)
    today_records = [r for r in attendance_list if r['date'] == today]
    for s in staff_list:
        cls = next((c for c in classes_list if c['id'] == s.get('class_id')), None)
        s['class_name'] = cls['name'] if cls else 'N/A'
        rec = next((r for r in today_records if r['student_id'] == s['id']), None)
        s['today_status'] = rec['status'] if rec else 'not_marked'
        s['today_time_in'] = rec.get('time_in', '') if rec else ''
        s['today_time_out'] = rec.get('time_out', '') if rec else ''
    return render_template('qr_attendance.html',
        classes=classes_list, staff=staff_list, today=today)

@app.route('/attendance/qr/ai-generate/<sid>')
@login_required
def ai_qr_generate(sid):
    """Generate an AI-branded QR code image for a staff member.
    Uses the AI-generated decorative frame + Python qrcode lib for a real scannable QR."""
    try:
        return _ai_qr_generate_inner(sid)
    except Exception as e:
        app.logger.error(f'AI QR generation error for {sid}: {e}')
        return jsonify({'error': 'QR generation failed: ' + str(e)}), 500

def _ai_qr_generate_inner(sid):
    staff_list = load_json(STAFF_FILE)
    member = next((s for s in staff_list if s['id'] == sid), None)
    if not member:
        return jsonify({'error': 'Staff not found'}), 404

    # Build the QR scan URL
    scan_url = request.host_url + 'attendance/qr/scan?sid=' + sid

    # Ensure output directory exists
    ai_qr_dir = os.path.join(os.path.dirname(__file__), 'static', 'ai_qr')
    os.makedirs(ai_qr_dir, exist_ok=True)
    output_path = os.path.join(ai_qr_dir, sid + '.png')

    # If already generated today, serve cached version
    if os.path.exists(output_path):
        mtime = datetime.fromtimestamp(os.path.getmtime(output_path))
        if mtime.date() == date.today():
            return send_file(output_path, mimetype='image/png')

    # Generate QR code with qrcode library
    # Scannability: use pure black fill (#000000) for maximum contrast against
    # the white QR background.  The frame background is cream (~237-249 RGB),
    # so a white underlay rectangle is painted before pasting the QR.
    # NEAREST resampling preserves sharp module edges (no anti-aliasing).
    # border=2 is the minimum quiet zone but the decorative frame border
    # provides an additional visual quiet zone, allowing larger data modules.
    try:
        qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
        qr.add_data(scan_url)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color='#000000', back_color='#ffffff')
        qr_img = qr_img.convert('RGBA')
    except Exception as e:
        app.logger.error(f'QR generation failed for {sid}: {e}')
        # Fallback: generate basic QR without custom colors
        qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
        qr.add_data(scan_url)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color='#000000', back_color='#ffffff')
        qr_img = qr_img.convert('RGBA')

    # Open the AI decorative frame
    frame_path = os.path.join(os.path.dirname(__file__), 'static', 'ai_qr_frame.png')
    if not os.path.exists(frame_path):
        qr_img.save(output_path)
        return send_file(output_path, mimetype='image/png')

    frame = PILImage.open(frame_path).convert('RGBA')
    frame_w, frame_h = frame.size

    # Detect the center QR-placeholder region by scanning outward from
    # center to find the bright-to-dark border transitions. This approach
    # reliably finds the inside edges of the dark frame border regardless of
    # outer decorative elements, gradients, or grain in the frame artwork.
    try:
        import numpy as np
        frame_arr = np.array(frame)  # RGBA
        r_ch = frame_arr[:, :, 0]
        g_ch = frame_arr[:, :, 1]
        b_ch = frame_arr[:, :, 2]
        cx, cy = frame_w // 2, frame_h // 2
        y_band = slice(frame_h // 3, 2 * frame_h // 3)
        x_band = slice(frame_w // 3, 2 * frame_w // 3)

        # Scan left from center: find the first column where the mean brightness
        # drops below 120 (dark border), then scan back right to find the
        # bright-to-dark transition (the inside edge of the border).
        left_edge_x = None
        for x in range(cx, 0, -1):
            mr = np.mean(r_ch[y_band, x])
            mg = np.mean(g_ch[y_band, x])
            mb = np.mean(b_ch[y_band, x])
            if mr < 120 and mg < 120 and mb < 120:
                for x2 in range(x + 1, cx):
                    mr2 = np.mean(r_ch[y_band, x2])
                    mg2 = np.mean(g_ch[y_band, x2])
                    mb2 = np.mean(b_ch[y_band, x2])
                    if mr2 > 150 and mg2 > 150 and mb2 > 150:
                        left_edge_x = x2
                        break
                break

        # Scan right from center
        right_edge_x = None
        for x in range(cx, frame_w):
            mr = np.mean(r_ch[y_band, x])
            mg = np.mean(g_ch[y_band, x])
            mb = np.mean(b_ch[y_band, x])
            if mr < 120 and mg < 120 and mb < 120:
                for x2 in range(x - 1, cx, -1):
                    mr2 = np.mean(r_ch[y_band, x2])
                    mg2 = np.mean(g_ch[y_band, x2])
                    mb2 = np.mean(b_ch[y_band, x2])
                    if mr2 > 150 and mg2 > 150 and mb2 > 150:
                        right_edge_x = x2
                        break
                break

        # Scan up from center
        top_edge_y = None
        for y in range(cy, 0, -1):
            mr = np.mean(r_ch[y, x_band])
            mg = np.mean(g_ch[y, x_band])
            mb = np.mean(b_ch[y, x_band])
            if mr < 120 and mg < 120 and mb < 120:
                for y2 in range(y + 1, cy):
                    mr2 = np.mean(r_ch[y2, x_band])
                    mg2 = np.mean(g_ch[y2, x_band])
                    mb2 = np.mean(b_ch[y2, x_band])
                    if mr2 > 150 and mg2 > 150 and mb2 > 150:
                        top_edge_y = y2
                        break
                break

        # Scan down from center
        bottom_edge_y = None
        for y in range(cy, frame_h):
            mr = np.mean(r_ch[y, x_band])
            mg = np.mean(g_ch[y, x_band])
            mb = np.mean(b_ch[y, x_band])
            if mr < 120 and mg < 120 and mb < 120:
                for y2 in range(y - 1, cy, -1):
                    mr2 = np.mean(r_ch[y2, x_band])
                    mg2 = np.mean(g_ch[y2, x_band])
                    mb2 = np.mean(b_ch[y2, x_band])
                    if mr2 > 150 and mg2 > 150 and mb2 > 150:
                        bottom_edge_y = y2
                        break
                break

        if all(v is not None for v in [left_edge_x, right_edge_x, top_edge_y, bottom_edge_y]):
            # Found the bright center region inside the dark borders.
            # Add a margin so QR clears any gold corner brackets / inner accents.
            margin = 15
            qr_w = right_edge_x - left_edge_x
            qr_h = bottom_edge_y - top_edge_y
            qr_target_size = min(qr_w, qr_h) - 2 * margin
            center_x = (left_edge_x + right_edge_x) // 2
            center_y = (top_edge_y + bottom_edge_y) // 2
            app.logger.info(
                f'AI QR: center region x={left_edge_x}-{right_edge_x}, '
                f'y={top_edge_y}-{bottom_edge_y}, '
                f'qr_size={qr_target_size}, center=({center_x},{center_y})'
            )
        else:
            # Hardcoded fallback based on ai_qr_frame.png (1024x1024) analysis:
            # Dark border at x=272-275 (L), x=748-751 (R), y=326-329 (T), y=799 (B)
            # Bright inside edges: x=276-747, y=331-797
            # Safe QR region with margin for gold brackets: x=295-735, y=345-785
            center_x = 515
            center_y = 564
            qr_target_size = 420
            app.logger.info('AI QR: using hardcoded fallback placement coordinates')
    except (ImportError, Exception) as e:
        app.logger.warning(f'AI QR detection failed ({e}), using defaults')
        # Hardcoded fallback for 1024x1024 frame
        center_x = 515
        center_y = 564
        qr_target_size = 420

    # NEAREST resampling preserves sharp black/white module edges.
    # LANCZOS creates anti-aliasing artifacts (gray pixels at module edges)
    # that can confuse QR scanners, especially at lower resolutions.
    qr_resized = qr_img.resize((qr_target_size, qr_target_size), PILImage.NEAREST)

    qr_x = center_x - qr_target_size // 2
    qr_y = center_y - qr_target_size // 2

    # Paint a solid white rectangle on the frame at the QR placement area
    # before pasting.  The frame center is cream (~237-249 RGB), not pure
    # white.  Without this underlay, the QR's white modules blend with the
    # cream background, reducing contrast with the dark modules.
    composite = frame.copy()
    from PIL import ImageDraw
    draw = ImageDraw.Draw(composite)
    draw.rectangle(
        [qr_x, qr_y, qr_x + qr_target_size - 1, qr_y + qr_target_size - 1],
        fill=(255, 255, 255, 255)
    )
    composite.paste(qr_resized, (qr_x, qr_y), qr_resized)

    composite = composite.convert('RGB')
    composite.save(output_path, 'PNG', optimize=True)
    return send_file(output_path, mimetype='image/png')


@app.route('/attendance/qr/image/<sid>')
def qr_image_fallback(sid):
    """Server-side QR image fallback — returns a simple scannable QR PNG.
    No login required so the client can use it as an <img> fallback when
    the qrcode.js CDN fails.  The QR encodes the scan URL for the given
    staff member, exactly the same data the client-side canvas would use."""
    staff_list = load_json(STAFF_FILE)
    member = next((s for s in staff_list if s['id'] == sid), None)
    if not member:
        return jsonify({'error': 'Staff not found'}), 404

    scan_url = request.host_url + 'attendance/qr/scan?sid=' + sid

    buf = io.BytesIO()
    qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M,
                        box_size=10, border=2)
    qr.add_data(scan_url)
    qr.make(fit=True)
    # Pure black fill for maximum scannability contrast
    qr_img = qr.make_image(fill_color='#000000', back_color='#ffffff')
    qr_img = qr_img.convert('RGB')
    qr_img.save(buf, 'PNG', optimize=True)
    buf.seek(0)
    return send_file(buf, mimetype='image/png')


@app.route('/attendance/qr/scan', methods=['GET', 'POST'])
def qr_scan_entry():
    """GET: Show QR landing page; POST: Legacy scanner API."""
    if request.method == 'GET':
        return _qr_scan_landing()
    return qr_scan_attendance_api()

def _qr_scan_landing():
    """Show landing page with staff name + check-in/check-out buttons.
    sid is passed as query param: /attendance/qr/scan?sid=S001"""
    sid = request.args.get('sid', '')
    if not sid:
        return render_template('qr_scan_landing.html', staff=None, error='No staff ID provided in QR code.')
    staff_list = load_json(STAFF_FILE)
    member = next((s for s in staff_list if s['id'] == sid), None)
    if not member:
        return render_template('qr_scan_landing.html', staff=None, error=f'Staff {sid} not found.')
    classes = load_json(CLASSES_FILE)
    cls = next((c for c in classes if c['id'] == member.get('class_id')), None)
    member['class_name'] = cls['name'] if cls else 'N/A'
    today = date.today().isoformat()
    attendance_list = load_json(ATTENDANCE_FILE)
    today_rec = next((r for r in attendance_list if r['student_id'] == sid and r['date'] == today), None)
    member['checked_in'] = bool(today_rec and today_rec.get('time_in') and not today_rec.get('time_out'))
    member['checked_out'] = bool(today_rec and today_rec.get('time_in') and today_rec.get('time_out'))
    member['not_marked'] = not today_rec or not today_rec.get('time_in')
    member['time_in'] = today_rec.get('time_in', '') if today_rec else ''
    member['time_out'] = today_rec.get('time_out', '') if today_rec else ''
    member['today'] = today
    # Pass geofence info for GPS display
    if cls and 'lat' in cls:
        member['centre_lat'] = cls['lat']
        member['centre_lon'] = cls['lon']
        member['centre_radius'] = cls.get('radius_meters', 200)
    loc_info = get_client_location_info()
    return render_template('qr_scan_landing.html', staff=member, error=None, client_ip=loc_info['ip_address'], client_location=f"{loc_info['ip_city']}, {loc_info['ip_region']}, {loc_info['ip_country']}")

def qr_scan_attendance_api():
    """POST API for camera scanner / manual code entry."""
    data = request.get_json()
    code = data.get('code', '')
    # Extract GPS data from client
    gps_lat = data.get('gps_lat', '')
    gps_lon = data.get('gps_lon', '')
    gps_accuracy = data.get('gps_accuracy', '')
    sid = None
    if code.startswith('http'):
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(code)
        qs = parse_qs(parsed.query)
        sid = qs.get('sid', [None])[0]
        if not sid:
            parts = parsed.path.rstrip('/').split('/')
            if parts and parts[-1] != 'scan':
                sid = parts[-1]
    elif code.startswith('TC:'):
        parts = code.split(':')
        if len(parts) >= 2:
            sid = parts[1]
    if not sid:
        return jsonify({'success': False, 'message': 'Invalid QR code format. Expected TC:S001:date or a scannable URL.'})
    today = date.today().isoformat()
    staff_list = load_json(STAFF_FILE)
    member = next((s for s in staff_list if s['id'] == sid), None)
    if not member:
        return jsonify({'success': False, 'message': 'Staff not found'})
    # ── GPS Geofence Check ──
    class_id = member.get('class_id', '')
    geo_ok, geo_dist, geo_centre, geo_clat, geo_clon, geo_radius = check_geofence(gps_lat, gps_lon, class_id)
    if geo_ok is False:
        # Staff is OUTSIDE the geofence — reject
        return jsonify({
            'success': False,
            'message': f'🚫 Geofence check failed: You are {geo_dist}m away from {geo_centre} (allowed: {geo_radius}m). Please check in from your Location.',
            'geofence_status': 'outside',
            'distance_m': geo_dist,
            'radius_m': geo_radius,
            'centre_name': geo_centre
        })
    geofence_status = 'verified' if geo_ok is True else 'no_gps'
    attendance_list = load_json(ATTENDANCE_FILE)
    existing = next((r for r in attendance_list if r['student_id'] == sid and r['date'] == today), None)
    now_time = datetime.now().strftime('%H:%M')
    loc_info = get_client_location_info()
    if existing:
        if existing.get('time_in') and not existing.get('time_out'):
            existing['time_out'] = now_time
            existing['status'] = 'present'
            existing['checkout_ip_address'] = loc_info['ip_address']
            existing['checkout_ip_city'] = loc_info['ip_city']
            existing['checkout_ip_region'] = loc_info['ip_region']
            existing['checkout_ip_country'] = loc_info['ip_country']
            existing['checkout_gps_lat'] = str(gps_lat)
            existing['checkout_gps_lon'] = str(gps_lon)
            existing['checkout_gps_accuracy'] = str(gps_accuracy)
            existing['checkout_geofence_status'] = geofence_status
            if geo_ok is True:
                existing['checkout_geofence_distance'] = geo_dist
            save_json(ATTENDANCE_FILE, attendance_list)
            geo_msg = f' (GPS: {geo_dist}m from centre)' if geo_ok is True else ' (GPS not verified)'
            return jsonify({'success': True, 'action': 'timeout', 'name': member['name'],
                'time_out': now_time, 'time_in': existing.get('time_in', ''), 'message': f'Time-out recorded for {member["name"]} at {now_time}{geo_msg}', 'geofence_status': geofence_status})
        elif existing.get('time_in') and existing.get('time_out'):
            return jsonify({'success': False, 'message': f'{member["name"]} already checked in and out today.'})
        else:
            existing['time_in'] = now_time
            existing['status'] = 'present'
            existing['ip_address'] = loc_info['ip_address']
            existing['ip_city'] = loc_info['ip_city']
            existing['ip_region'] = loc_info['ip_region']
            existing['ip_country'] = loc_info['ip_country']
            existing['gps_lat'] = str(gps_lat)
            existing['gps_lon'] = str(gps_lon)
            existing['gps_accuracy'] = str(gps_accuracy)
            existing['geofence_status'] = geofence_status
            if geo_ok is True:
                existing['geofence_distance'] = geo_dist
            save_json(ATTENDANCE_FILE, attendance_list)
            geo_msg = f' (GPS: {geo_dist}m from centre)' if geo_ok is True else ' (GPS not verified)'
            return jsonify({'success': True, 'action': 'timein', 'name': member['name'],
                'time_in': now_time, 'message': f'Time-in recorded for {member["name"]} at {now_time}{geo_msg}', 'geofence_status': geofence_status})
    else:
        new_rec = {
            'id': generate_id('A', attendance_list),
            'student_id': sid,
            'class_id': member.get('class_id', ''),
            'date': today,
            'status': 'present',
            'time_in': now_time,
            'time_out': '',
            'marked_by': request.cookies.get('user_name', 'Unknown'),
            'marked_at': datetime.now().isoformat(),
            'ip_address': loc_info['ip_address'],
            'ip_city': loc_info['ip_city'],
            'ip_region': loc_info['ip_region'],
            'ip_country': loc_info['ip_country'],
            'gps_lat': str(gps_lat),
            'gps_lon': str(gps_lon),
            'gps_accuracy': str(gps_accuracy),
            'geofence_status': geofence_status
        }
        if geo_ok is True:
            new_rec['geofence_distance'] = geo_dist
        attendance_list.append(new_rec)
        save_json(ATTENDANCE_FILE, attendance_list)
        geo_msg = f' (GPS: {geo_dist}m from centre)' if geo_ok is True else ' (GPS not verified)'
        return jsonify({'success': True, 'action': 'timein', 'name': member['name'],
            'time_in': now_time, 'message': f'Time-in recorded for {member["name"]} at {now_time}{geo_msg}', 'geofence_status': geofence_status})

@app.route('/attendance/qr/checkin', methods=['POST'])
def qr_checkin():
    """Check in a staff member via the QR landing page."""
    sid = request.form.get('sid', '')
    # Extract GPS data from hidden form fields
    gps_lat = request.form.get('gps_lat', '')
    gps_lon = request.form.get('gps_lon', '')
    gps_accuracy = request.form.get('gps_accuracy', '')
    if not sid:
        flash('No staff ID provided.', 'danger')
        return redirect(url_for('qr_scan_entry', sid=''))
    today = date.today().isoformat()
    now_time = datetime.now().strftime('%H:%M')
    staff_list = load_json(STAFF_FILE)
    member = next((s for s in staff_list if s['id'] == sid), None)
    if not member:
        flash('Staff not found.', 'danger')
        return redirect(url_for('qr_scan_entry', sid=''))
    # ── GPS Geofence Check ──
    class_id = member.get('class_id', '')
    geo_ok, geo_dist, geo_centre, geo_clat, geo_clon, geo_radius = check_geofence(gps_lat, gps_lon, class_id)
    if geo_ok is False:
        # Staff is OUTSIDE the geofence — reject check-in
        flash(f'🚫 Geofence check failed: You are {geo_dist}m away from {geo_centre} (allowed: {geo_radius}m). Please check in from your Location.', 'danger')
        return redirect(url_for('qr_scan_entry', sid=sid))
    geofence_status = 'verified' if geo_ok is True else 'no_gps'
    attendance_list = load_json(ATTENDANCE_FILE)
    existing = next((r for r in attendance_list if r['student_id'] == sid and r['date'] == today), None)
    if existing and existing.get('time_in'):
        flash(f'{member["name"]} is already checked in at {existing.get("time_in", "?")}.', 'warning')
        return redirect(url_for('qr_scan_entry', sid=sid))
    loc_info = get_client_location_info()
    if existing:
        existing['time_in'] = now_time
        existing['status'] = 'present'
        existing['marked_by'] = request.cookies.get('user_name', 'QR Check-in')
        existing['marked_at'] = datetime.now().isoformat()
        existing['ip_address'] = loc_info['ip_address']
        existing['ip_city'] = loc_info['ip_city']
        existing['ip_region'] = loc_info['ip_region']
        existing['ip_country'] = loc_info['ip_country']
        existing['gps_lat'] = str(gps_lat)
        existing['gps_lon'] = str(gps_lon)
        existing['gps_accuracy'] = str(gps_accuracy)
        existing['geofence_status'] = geofence_status
        if geo_ok is True:
            existing['geofence_distance'] = geo_dist
    else:
        new_rec = {
            'id': generate_id('A', attendance_list),
            'student_id': sid,
            'class_id': member.get('class_id', ''),
            'date': today,
            'status': 'present',
            'time_in': now_time,
            'time_out': '',
            'marked_by': request.cookies.get('user_name', 'QR Check-in'),
            'marked_at': datetime.now().isoformat(),
            'ip_address': loc_info['ip_address'],
            'ip_city': loc_info['ip_city'],
            'ip_region': loc_info['ip_region'],
            'ip_country': loc_info['ip_country'],
            'gps_lat': str(gps_lat),
            'gps_lon': str(gps_lon),
            'gps_accuracy': str(gps_accuracy),
            'geofence_status': geofence_status
        }
        if geo_ok is True:
            new_rec['geofence_distance'] = geo_dist
        attendance_list.append(new_rec)
    save_json(ATTENDANCE_FILE, attendance_list)
    geo_msg = f' GPS verified ({geo_dist}m from centre).' if geo_ok is True else ' GPS not verified.'
    flash(f'✅ Check-in recorded for {member["name"]} at {now_time}.{geo_msg}', 'success')
    return redirect(url_for('qr_scan_entry', sid=sid))

@app.route('/attendance/qr/checkout', methods=['POST'])
def qr_checkout():
    """Check out a staff member via the QR landing page."""
    sid = request.form.get('sid', '')
    # Extract GPS data from hidden form fields
    gps_lat = request.form.get('gps_lat', '')
    gps_lon = request.form.get('gps_lon', '')
    gps_accuracy = request.form.get('gps_accuracy', '')
    if not sid:
        flash('No staff ID provided.', 'danger')
        return redirect(url_for('qr_scan_entry', sid=''))
    today = date.today().isoformat()
    now_time = datetime.now().strftime('%H:%M')
    staff_list = load_json(STAFF_FILE)
    member = next((s for s in staff_list if s['id'] == sid), None)
    if not member:
        flash('Staff not found.', 'danger')
        return redirect(url_for('qr_scan_entry', sid=''))
    # ── GPS Geofence Check ──
    class_id = member.get('class_id', '')
    geo_ok, geo_dist, geo_centre, geo_clat, geo_clon, geo_radius = check_geofence(gps_lat, gps_lon, class_id)
    if geo_ok is False:
        # Staff is OUTSIDE the geofence — reject check-out
        flash(f'🚫 Geofence check failed: You are {geo_dist}m away from {geo_centre} (allowed: {geo_radius}m). Please check out from your Location.', 'danger')
        return redirect(url_for('qr_scan_entry', sid=sid))
    geofence_status = 'verified' if geo_ok is True else 'no_gps'
    attendance_list = load_json(ATTENDANCE_FILE)
    existing = next((r for r in attendance_list if r['student_id'] == sid and r['date'] == today), None)
    if not existing or not existing.get('time_in'):
        flash(f'{member["name"]} has not checked in yet. Please check in first.', 'warning')
        return redirect(url_for('qr_scan_entry', sid=sid))
    if existing.get('time_out'):
        flash(f'{member["name"]} already checked out at {existing.get("time_out", "?")}.', 'warning')
        return redirect(url_for('qr_scan_entry', sid=sid))
    existing['time_out'] = now_time
    existing['status'] = 'present'
    existing['marked_by'] = request.cookies.get('user_name', 'QR Check-out')
    existing['marked_at'] = datetime.now().isoformat()
    loc_info = get_client_location_info()
    existing['checkout_ip_address'] = loc_info['ip_address']
    existing['checkout_ip_city'] = loc_info['ip_city']
    existing['checkout_ip_region'] = loc_info['ip_region']
    existing['checkout_ip_country'] = loc_info['ip_country']
    existing['checkout_gps_lat'] = str(gps_lat)
    existing['checkout_gps_lon'] = str(gps_lon)
    existing['checkout_gps_accuracy'] = str(gps_accuracy)
    existing['checkout_geofence_status'] = geofence_status
    if geo_ok is True:
        existing['checkout_geofence_distance'] = geo_dist
    save_json(ATTENDANCE_FILE, attendance_list)
    geo_msg = f' GPS verified ({geo_dist}m from centre).' if geo_ok is True else ' GPS not verified.'
    flash(f'✅ Check-out recorded for {member["name"]} at {now_time}.{geo_msg}', 'success')
    return redirect(url_for('qr_scan_entry', sid=sid))

# ─── Time Records ───
@app.route('/attendance/time-records')
@login_required
def time_records():
    attendance_list = load_json(ATTENDANCE_FILE)
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    today = date.today().isoformat()
    class_id = request.args.get('class_id', '')
    rec_date = request.args.get('date', today)
    filtered = [r for r in attendance_list if r.get('time_in') and r['date'] == rec_date]
    if class_id:
        filtered = [r for r in filtered if r['class_id'] == class_id]
    for r in filtered:
        member = next((s for s in staff_list if s['id'] == r['student_id']), None)
        cls = next((c for c in classes_list if c['id'] == r['class_id']), None)
        r['student_name'] = member['name'] if member else 'Unknown'
        r['roll_no'] = member.get('roll_no', '') if member else ''
        r['class_name'] = cls['name'] if cls else 'N/A'
    return render_template('time_records.html',
        records=filtered, classes=classes_list,
        selected_class=class_id, selected_date=rec_date, today=today)

@app.route('/attendance/time-records/export')
@login_required
def export_time_records():
    fmt = request.args.get('format', 'csv')
    attendance_list = load_json(ATTENDANCE_FILE)
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    class_id = request.args.get('class_id', '')
    rec_date = request.args.get('date', date.today().isoformat())
    filtered = [r for r in attendance_list if r.get('time_in') and r['date'] == rec_date]
    if class_id:
        filtered = [r for r in filtered if r['class_id'] == class_id]
    headers = ['Date', 'Location', 'Staff ID', 'Staff Name', 'Status', 'Time In', 'Check-in Location', 'Time Out', 'Check-out Location', 'Hours']
    rows = []
    for r in filtered:
        member = next((s for s in staff_list if s['id'] == r['student_id']), None)
        cls = next((c for c in classes_list if c['id'] == r['class_id']), None)
        ti = r.get('time_in', '')
        to = r.get('time_out', '')
        checkin_loc = f"{r.get('ip_city', '')}, {r.get('ip_country', '')}".strip(', ') or r.get('ip_address', '')
        checkout_loc = f"{r.get('checkout_ip_city', '')}, {r.get('checkout_ip_country', '')}".strip(', ') or r.get('checkout_ip_address', '')
        hours = ''
        if ti and to:
            try:
                h = (datetime.strptime(to, '%H:%M') - datetime.strptime(ti, '%H:%M')).total_seconds() / 3600
                hours = f'{h:.1f}'
            except:
                pass
        rows.append([
            r['date'],
            cls['name'] if cls else 'N/A',
            member.get('roll_no', '') if member else '',
            member['name'] if member else 'Unknown',
            r['status'].capitalize(),
            ti, checkin_loc, to, checkout_loc, hours
        ])
    if fmt == 'excel':
        wb = make_excel(headers, rows, 'Time Records')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'time_records_{rec_date}.xlsx')
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv', as_attachment=True, download_name=f'time_records_{rec_date}.csv')

@app.route('/attendance/records')
@login_required
def attendance_records():
    attendance_list = load_json(ATTENDANCE_FILE)
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    for r in attendance_list:
        member = next((s for s in staff_list if s['id'] == r['student_id']), None)
        cls = next((c for c in classes_list if c['id'] == r['class_id']), None)
        r['student_name'] = member['name'] if member else 'Unknown'
        r['roll_no'] = member.get('roll_no', '') if member else ''
        r['class_name'] = cls['name'] if cls else 'N/A'
    attendance_list.sort(key=lambda x: x.get('date', ''), reverse=True)
    return render_template('attendance_records.html',
        records=attendance_list, classes=classes_list)

@app.route('/attendance/records/data')
@login_required
def attendance_records_data():
    attendance_list = load_json(ATTENDANCE_FILE)
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    class_id = request.args.get('class_id', '')
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    if class_id:
        attendance_list = [r for r in attendance_list if r['class_id'] == class_id]
    if date_from:
        attendance_list = [r for r in attendance_list if r['date'] >= date_from]
    if date_to:
        attendance_list = [r for r in attendance_list if r['date'] <= date_to]
    result = []
    for r in attendance_list:
        member = next((s for s in staff_list if s['id'] == r['student_id']), None)
        cls = next((c for c in classes_list if c['id'] == r['class_id']), None)
        result.append({
            'id': r['id'],
            'date': r['date'],
            'roll_no': member.get('roll_no', '') if member else '',
            'student_name': member['name'] if member else 'Unknown',
            'class_name': cls['name'] if cls else 'N/A',
            'status': r['status'],
            'time_in': r.get('time_in', ''),
            'time_out': r.get('time_out', ''),
            'ip_address': r.get('ip_address', ''),
            'ip_city': r.get('ip_city', ''),
            'ip_region': r.get('ip_region', ''),
            'ip_country': r.get('ip_country', ''),
            'checkout_ip_address': r.get('checkout_ip_address', ''),
            'checkout_ip_city': r.get('checkout_ip_city', ''),
            'checkout_ip_region': r.get('checkout_ip_region', ''),
            'checkout_ip_country': r.get('checkout_ip_country', '')
        })
    return jsonify(result)

@app.route('/attendance/records/export')
@login_required
def export_attendance_records():
    fmt = request.args.get('format', 'csv')
    attendance_list = load_json(ATTENDANCE_FILE)
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    class_id = request.args.get('class_id', '')
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    if class_id:
        attendance_list = [r for r in attendance_list if r['class_id'] == class_id]
    if date_from:
        attendance_list = [r for r in attendance_list if r['date'] >= date_from]
    if date_to:
        attendance_list = [r for r in attendance_list if r['date'] <= date_to]
    headers = ['Date', 'Location', 'Staff ID', 'Staff Name', 'Status', 'Time In', 'Check-in Location', 'Time Out', 'Check-out Location']
    rows = []
    for r in attendance_list:
        member = next((s for s in staff_list if s['id'] == r['student_id']), None)
        cls = next((c for c in classes_list if c['id'] == r['class_id']), None)
        checkin_loc = f"{r.get('ip_city', '')}, {r.get('ip_country', '')}".strip(', ') or r.get('ip_address', '')
        checkout_loc = f"{r.get('checkout_ip_city', '')}, {r.get('checkout_ip_country', '')}".strip(', ') or r.get('checkout_ip_address', '')
        rows.append([
            r['date'],
            cls['name'] if cls else 'N/A',
            member.get('roll_no', '') if member else '',
            member['name'] if member else 'Unknown',
            r['status'].capitalize(),
            r.get('time_in', ''),
            checkin_loc,
            r.get('time_out', ''),
            checkout_loc
        ])
    if fmt == 'excel':
        wb = make_excel(headers, rows, 'Attendance Records')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='attendance_records.xlsx')
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv', as_attachment=True, download_name='attendance_records.csv')

# ─── Admin Attendance Edit/Delete ───
@app.route('/attendance/edit/<att_id>', methods=['GET'])
@login_required
@admin_required
def get_attendance_edit(att_id):
    attendance_list = load_json(ATTENDANCE_FILE)
    record = next((r for r in attendance_list if r['id'] == att_id), None)
    if not record:
        return jsonify({'success': False, 'message': 'Record not found'}), 404
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    member = next((s for s in staff_list if s['id'] == record['student_id']), None)
    cls = next((c for c in classes_list if c['id'] == record['class_id']), None)
    return jsonify({
        'success': True,
        'id': record['id'],
        'student_id': record['student_id'],
        'class_id': record['class_id'],
        'date': record['date'],
        'status': record['status'],
        'time_in': record.get('time_in', ''),
        'time_out': record.get('time_out', ''),
        'student_name': member['name'] if member else 'Unknown',
        'class_name': cls['name'] if cls else 'N/A'
    })

@app.route('/attendance/edit/<att_id>', methods=['POST'])
@login_required
@admin_required
def save_attendance_edit(att_id):
    attendance_list = load_json(ATTENDANCE_FILE)
    record = next((r for r in attendance_list if r['id'] == att_id), None)
    if not record:
        return jsonify({'success': False, 'message': 'Record not found'}), 404
    data = request.get_json()
    if data.get('status'):
        record['status'] = data['status']
    if 'time_in' in data:
        record['time_in'] = data['time_in']
    if 'time_out' in data:
        record['time_out'] = data['time_out']
    if data.get('date'):
        record['date'] = data['date']
    save_json(ATTENDANCE_FILE, attendance_list)
    return jsonify({'success': True, 'message': 'Record updated successfully'})

@app.route('/attendance/delete/<att_id>', methods=['POST'])
@login_required
@admin_required
def delete_attendance(att_id):
    attendance_list = load_json(ATTENDANCE_FILE)
    record = next((r for r in attendance_list if r['id'] == att_id), None)
    if not record:
        return jsonify({'success': False, 'message': 'Record not found'}), 404
    attendance_list = [r for r in attendance_list if r['id'] != att_id]
    save_json(ATTENDANCE_FILE, attendance_list)
    return jsonify({'success': True, 'message': 'Record deleted successfully'})

# ─── Reports ───
@app.route('/reports')
@login_required
def reports():
    staff_list = load_json(STAFF_FILE)
    classes_list = load_json(CLASSES_FILE)
    attendance_list = load_json(ATTENDANCE_FILE)

    centre_reports = []
    for c in classes_list:
        c_staff = [s for s in staff_list if s.get('class_id') == c['id'] and s.get('status') == 'active']
        c_records = [r for r in attendance_list if r['class_id'] == c['id']]
        total = len(c_records)
        present = sum(1 for r in c_records if r['status'] == 'present')
        absent = sum(1 for r in c_records if r['status'] == 'absent')
        rate = round(present / total * 100, 1) if total > 0 else 0
        centre_reports.append({
            'name': c['name'], 'total': len(c_staff),
            'present': present, 'absent': absent, 'rate': rate
        })

    weekly_data = []
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        day_records = [r for r in attendance_list if r['date'] == d]
        weekly_data.append({
            'date': d,
            'present': sum(1 for r in day_records if r['status'] == 'present'),
            'absent': sum(1 for r in day_records if r['status'] == 'absent')
        })

    monthly_data = []
    for i in range(5, -1, -1):
        month_date = (date.today().replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_str = month_date.strftime('%Y-%m')
        month_records = [r for r in attendance_list if r['date'].startswith(month_str)]
        total = len(month_records)
        present = sum(1 for r in month_records if r['status'] == 'present')
        rate = round(present / total * 100, 1) if total > 0 else 0
        monthly_data.append({'month': month_str, 'rate': rate})

    return render_template('reports.html',
        centre_reports=centre_reports,
        weekly_data=weekly_data, monthly_data=monthly_data,
        staff=staff_list, classes=classes_list)

@app.route('/reports/staff/<sid>')
@login_required
def staff_report(sid):
    attendance_list = load_json(ATTENDANCE_FILE)
    month = request.args.get('month', '')
    records = [r for r in attendance_list if r['student_id'] == sid]
    if month:
        records = [r for r in records if r['date'].startswith(month)]
    total = len(records)
    present = sum(1 for r in records if r['status'] == 'present')
    rate = round(present / total * 100, 1) if total > 0 else 0
    return jsonify({'records': [{'date': r['date'], 'status': r['status'], 'time_in': r.get('time_in',''), 'time_out': r.get('time_out','')} for r in records], 'rate': rate})

@app.route('/reports/staff/<sid>/export')
@login_required
def export_staff_report(sid):
    fmt = request.args.get('format', 'csv')
    attendance_list = load_json(ATTENDANCE_FILE)
    staff_list = load_json(STAFF_FILE)
    month = request.args.get('month', '')
    member = next((s for s in staff_list if s['id'] == sid), None)
    records = [r for r in attendance_list if r['student_id'] == sid]
    if month:
        records = [r for r in records if r['date'].startswith(month)]
    name = member['name'] if member else 'staff'
    headers = ['Date', 'Status', 'Time In', 'Time Out']
    rows = [[r['date'], r['status'].capitalize(), r.get('time_in',''), r.get('time_out','')] for r in records]
    if fmt == 'excel':
        wb = make_excel(headers, rows, name[:31])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'{name}_report.xlsx')
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv', as_attachment=True, download_name=f'{name}_report.csv')

# ─── API for dashboard chart ───
@app.route('/api/attendance-chart')
@login_required
def attendance_chart():
    attendance_list = load_json(ATTENDANCE_FILE)
    last_7 = []
    today = date.today()
    for i in range(6, -1, -1):
        d = (today - timedelta(days=i)).isoformat()
        records = [r for r in attendance_list if r['date'] == d]
        last_7.append({
            'date': d,
            'present': sum(1 for r in records if r['status'] == 'present'),
            'absent': sum(1 for r in records if r['status'] == 'absent'),
            'late': sum(1 for r in records if r['status'] == 'late')
        })
    return jsonify(last_7)

@app.route('/api/check-telegram-link', methods=['POST'])
def api_check_telegram_link():
    """Check if a user's Telegram account has been linked (called from link_telegram page)."""
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    if not username:
        return jsonify({'linked': False, 'error': 'Username required'}), 400

    users = load_json(USERS_FILE)
    user = next((u for u in users if u.get('username', '').lower() == username.lower()), None)
    if not user:
        user = next((u for u in users if u.get('email', '').lower() == username.lower()), None)
    if not user:
        return jsonify({'linked': False, 'error': 'User not found'}), 404

    linked = bool(user.get('telegram_chat_id', ''))
    return jsonify({'linked': linked, 'username': user.get('username', '')})


@app.route('/api/classes')
@login_required
def api_classes():
    return jsonify(load_json(CLASSES_FILE))

# ─── Google Sign-In ───
def _get_google_oauth_urls():
    import urllib.request
    try:
        with urllib.request.urlopen(GOOGLE_DISCOVERY_URL, timeout=5) as resp:
            doc = json.loads(resp.read())
        return doc.get('authorization_endpoint'), doc.get('token_endpoint'), doc.get('userinfo_endpoint')
    except Exception:
        return None, None, None

@app.route('/auth/google')
def auth_google():
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        flash('Google Sign-In is not configured. Please use username/password login.', 'warning')
        return redirect(url_for('login'))
    auth_ep, _, _ = _get_google_oauth_urls()
    if not auth_ep:
        flash('Could not reach Google authentication service. Please try again later.', 'danger')
        return redirect(url_for('login'))
    redirect_uri = request.host_url.rstrip('/') + url_for('auth_google_callback')
    params = {
        'client_id': GOOGLE_CLIENT_ID,
        'redirect_uri': redirect_uri,
        'response_type': 'code',
        'scope': 'openid email profile',
        'access_type': 'offline',
        'prompt': 'consent',
    }
    return redirect(f'{auth_ep}?{urlencode(params)}')

@app.route('/auth/google/callback')
def auth_google_callback():
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        flash('Google Sign-In is not configured.', 'warning')
        return redirect(url_for('login'))
    code = request.args.get('code')
    error = request.args.get('error')
    if error or not code:
        flash(f'Google Sign-In was cancelled or failed: {error or "no code"}', 'danger')
        return redirect(url_for('login'))
    _, token_ep, userinfo_ep = _get_google_oauth_urls()
    if not token_ep:
        flash('Could not reach Google authentication service.', 'danger')
        return redirect(url_for('login'))
    redirect_uri = request.host_url.rstrip('/') + url_for('auth_google_callback')
    import urllib.request
    token_payload = urlencode({
        'code': code,
        'client_id': GOOGLE_CLIENT_ID,
        'client_secret': GOOGLE_CLIENT_SECRET,
        'redirect_uri': redirect_uri,
        'grant_type': 'authorization_code',
    }).encode()
    try:
        req = urllib.request.Request(token_ep, data=token_payload, method='POST')
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        with urllib.request.urlopen(req, timeout=10) as resp:
            token_data = json.loads(resp.read())
        access_token = token_data.get('access_token')
        if not access_token:
            flash('Failed to obtain access token from Google.', 'danger')
            return redirect(url_for('login'))
    except Exception as e:
        flash(f'Error exchanging token with Google: {e}', 'danger')
        return redirect(url_for('login'))
    try:
        req = urllib.request.Request(userinfo_ep, method='GET')
        req.add_header('Authorization', f'Bearer {access_token}')
        with urllib.request.urlopen(req, timeout=10) as resp:
            guser = json.loads(resp.read())
    except Exception as e:
        flash(f'Error fetching Google profile: {e}', 'danger')
        return redirect(url_for('login'))
    google_email = guser.get('email', '')
    google_name = guser.get('name', guser.get('given_name', 'Google User'))
    google_sub = guser.get('sub', '')
    if not google_email:
        flash('Google account did not provide an email address.', 'danger')
        return redirect(url_for('login'))
    users = load_json(USERS_FILE)
    user = next((u for u in users if u.get('google_email', '').lower() == google_email.lower() or u.get('username', '').lower() == google_email.lower()), None)
    if not user:
        user = {
            'id': generate_id('U', users),
            'username': google_email,
            'password': hashlib.sha256(os.urandom(32)).hexdigest()[:16],
            'role': 'staff',
            'name': google_name,
            'phone': '',
            'email': google_email,
            'google_email': google_email,
            'google_sub': google_sub,
        }
        users.append(user)
        save_json(USERS_FILE, users)
    else:
        if not user.get('google_email'):
            user['google_email'] = google_email
            user['google_sub'] = google_sub
            save_json(USERS_FILE, users)
    resp = make_response(redirect(url_for('dashboard')))
    resp.set_cookie('user_id', user['id'], max_age=86400, httponly=True)
    resp.set_cookie('user_role', user['role'], max_age=86400, httponly=True)
    resp.set_cookie('user_name', user['name'], max_age=86400, httponly=False)
    flash(f'Welcome, {user["name"]}! Signed in via Google.', 'success')
    return resp

if __name__ == '__main__':
    debug = os.environ.get('FLASK_DEBUG', '1') == '1'
    app.run(debug=debug, host='0.0.0.0', port=int(os.environ.get('PORT', '5000')))
